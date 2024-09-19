import openpyxl
import random

# Open the existing Excel file
file_path = 'data.xlsx'  # Your file path
wb = openpyxl.load_workbook("E:\Projects\project04\dataset\data.xlsx")
sheet = wb.active  # Select the active sheet

# Optional: If you want to clear previous data but keep the headers, clear from row 2 onwards.
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        cell.value = None  # Clear cell value

# List of Sri Lankan districts (Add more if needed)
districts = [
    'Colombo', 'Gampaha', 'Kalutara', 'Kandy', 'Matale', 'Nuwara Eliya', 'Galle', 'Matara', 'Hambantota',
    'Jaffna', 'Kilinochchi', 'Mannar', 'Mullaitivu', 'Vavuniya', 'Trincomalee', 'Batticaloa', 'Ampara', 'Badulla',
    'Monaragala', 'Ratnapura', 'Kegalle', 'Puttalam', 'Kurunegala', 'Anuradhapura', 'Polonnaruwa'
]

# Define candidate mapping based on random number
candidate_map = {
    1: 'Anura Kumara',
    2: 'Sajith Premadasa',
    3: 'Ranil Wickremasinghe',
    4: 'Other'
}

# Define the number of voters (rows) you want to generate
number_of_voters = 5000  # You can change this to your desired number of rows

# Fill the columns with data
for row in range(2, number_of_voters + 2):  # Starting from row 2 (row 1 is for headers)
    
    # Counter column (Column A)
    sheet[f'A{row}'] = row - 1  # Sequential counter starting from 1
    
    # Age column (Column B)
    sheet[f'B{row}'] = random.randint(18, 100)  # Random age between 18 and 100
    
    # District column (Column C)
    sheet[f'C{row}'] = random.choice(districts)  # Random district from the list
    
    # Voting Number column (Column D)
    voting_number = random.randint(1, 4)  # Random number between 1 and 4 for candidate
    sheet[f'D{row}'] = voting_number
    
    # Candidate column (Column E) - Based on the voting number
    sheet[f'E{row}'] = candidate_map[voting_number]  # Assign candidate based on the voting number

# Save the updated workbook (overwrite the same file)
wb.save("E:\Projects\project04\dataset\data.xlsx")

print(f"Excel file '{file_path}' updated successfully!")